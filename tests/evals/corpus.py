"""Typed reader for the canonical Excel golden dataset."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tests.evals import config
from tests.evals.models import CorpusResult, EvalCase


class ExcelCorpusReader:
    def __init__(
        self,
        dataset_path: Path = config.DATASET_PATH,
        invoice_dir: Path = config.INVOICE_DIR,
    ) -> None:
        self._dataset_path = dataset_path
        self._invoice_dir = invoice_dir

    def read(self) -> CorpusResult:
        workbook = load_workbook(self._dataset_path, read_only=True, data_only=True)
        sheet = workbook["Sheet1"]
        rows = sheet.iter_rows(values_only=True)
        headers = [
            str(value).strip() if value is not None else "" for value in next(rows)
        ]
        result = CorpusResult()
        for row_number, values in enumerate(rows, start=2):
            result.total_rows += 1
            row = dict(zip(headers, values))
            eval_id = str(row.get("Eval ID") or "").strip()
            source = str(row.get("File Path") or "").strip()
            if not source:
                result.skipped_rows += 1
                continue
            pdf_path = self._resolve_source(source)
            if pdf_path is None:
                result.skipped_rows += 1
                result.issues.append(f"{eval_id}: source not found: {source}")
                continue
            expected = {
                field: row.get(header)
                for header, field in config.HEADER_TO_FIELD.items()
            }
            template_group = str(row.get("Template Group") or "").strip() or None
            result.cases.append(
                EvalCase(
                    eval_id=eval_id,
                    row_number=row_number,
                    pdf_path=pdf_path,
                    expected=expected,
                    template_group=template_group,
                )
            )
            result.executable_rows += 1
        workbook.close()
        return result

    def _resolve_source(self, source: str) -> Path | None:
        exact = self._invoice_dir / source
        if exact.is_file():
            return exact
        if not Path(source).suffix:
            with_pdf = self._invoice_dir / f"{source}.pdf"
            if with_pdf.is_file():
                return with_pdf
        matches = list(self._invoice_dir.glob(f"{source}*"))
        files = [path for path in matches if path.is_file()]
        return files[0] if len(files) == 1 else None


def workbook_value(value: Any) -> Any:
    """Kept as an explicit boundary for future workbook schema migrations."""
    return value
